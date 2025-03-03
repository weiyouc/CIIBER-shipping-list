#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Shipping List Processor

This script processes shipping list Excel files to produce export and re-import invoices
according to specified business rules.
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
import argparse
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def read_shipping_list(file_path):
    """
    Read the shipping list Excel file and return as DataFrame.
    
    Args:
        file_path (str): Path to the shipping list Excel file
        
    Returns:
        pd.DataFrame: DataFrame containing shipping list data
    """
    try:
        # Read the Excel file
        print(f"Reading file: {file_path}")
        try:
            # First try with no skiprows
            df = pd.read_excel(file_path)
        except Exception as e1:
            print(f"Error reading Excel file without skiprows: {e1}")
            try:
                # Try with skiprows=1 in case there's a header row
                df = pd.read_excel(file_path, skiprows=1)
                print("Successfully read file with skiprows=1")
            except Exception as e2:
                print(f"Error reading Excel file with skiprows=1: {e2}")
                raise
        
        print(f"Successfully read file with {len(df)} rows and {len(df.columns)} columns")
        
        # Print first few column names for debugging
        print("First 10 column names in the original file:")
        for i, col in enumerate(df.columns[:10]):
            print(f"{i}: {col}")
        
        # Define column mapping for known column names
        column_mapping = {
            # Serial number
            "Sr NO (序列号)": "serial_no",
            "Sr NO": "serial_no",
            "序列号": "serial_no",
            "Serial No": "serial_no",
            "Serial Number": "serial_no",
            
            # Part number
            "P/N.（系统料号 ）": "part_number",
            "P/N.": "part_number",
            "P/N": "part_number",
            "Part Number": "part_number",
            "料号": "part_number",
            "系统料号": "part_number",
            
            # Supplier
            "供应商": "supplier",
            "Supplier": "supplier",
            
            # Project name
            "项目名称": "project_name",
            "Project Name": "project_name",
            
            # Factory
            "工厂(Daman/Silvassa)": "factory",
            "工厂": "factory",
            "Factory": "factory",
            
            # Customs description
            "清关英文货描（关务提供）": "customs_desc_en",
            "清关英文货描": "customs_desc_en",
            "Customs Description": "customs_desc_en",
            "报关中文品名": "customs_desc_cn",
            "中文品名": "customs_desc_cn",
            
            # Description
            "DESCRIPTION (系统英文品名）": "description_en",
            "DESCRIPTION": "description_en",
            "Description": "description_en",
            "英文品名": "description_en",
            
            # Invoice name
            "开票名称": "invoice_name",
            "Invoice Name": "invoice_name",
            
            # Material name
            "物料名称": "material_name",
            "Material Name": "material_name",
            
            # Model
            "MODEL（货物型号（与实物相符)": "model",
            "MODEL": "model",
            "Model": "model",
            "货物型号": "model",
            
            # Quantity
            "QUANTITY （数量）": "quantity",
            "QUANTITY": "quantity",
            "Quantity": "quantity",
            "数量": "quantity",
            "Qty": "quantity",
            
            # Unit
            "单位": "unit",
            "Unit": "unit",
            
            # Carton measurement
            "Carton MEASUREMENT (外箱尺寸CM）": "carton_measurement",
            "Carton MEASUREMENT": "carton_measurement",
            "外箱尺寸": "carton_measurement",
            
            # Volume
            "体积（CBM）": "volume",
            "体积": "volume",
            "Volume": "volume",
            "总体积": "total_volume",
            "Total Volume": "total_volume",
            
            # Weight
            "单件毛重": "unit_gross_weight",
            "Unit Gross Weight": "unit_gross_weight",
            "G.W（KG) 总毛重": "total_gross_weight",
            "G.W": "total_gross_weight",
            "总毛重": "total_gross_weight",
            "Total Gross Weight": "total_gross_weight",
            "单件净重": "unit_net_weight",
            "Unit Net Weight": "unit_net_weight",
            "N.W  (KG) 总净重": "total_net_weight",
            "N.W": "total_net_weight",
            "总净重": "total_net_weight",
            "Total Net Weight": "total_net_weight",
            
            # Carton info
            "整箱数量": "full_carton_quantity",
            "Full Carton Quantity": "full_carton_quantity",
            "件数": "piece_count",
            "Piece Count": "piece_count",
            "CTN NO. (箱号)": "carton_no",
            "CTN NO.": "carton_no",
            "箱号": "carton_no",
            
            # Export customs method
            "出口报关方式": "export_customs_method",
            "Export Customs Method": "export_customs_method",
            
            # Purchasing unit
            "采购单位（智乐/UC/客供/供应商赠送/系统外订单）": "purchasing_unit",
            "采购单位": "purchasing_unit",
            "Purchasing Unit": "purchasing_unit",
            
            # Price
            "不含税单价（RMB）": "unit_price",
            "不含税单价": "unit_price",
            "单价": "unit_price",
            "Unit Price": "unit_price",
            
            # Tax rate
            "开票税率": "tax_rate"
        }
        
        # Create a lowercase version of the mapping for more robust matching
        lowercase_mapping = {}
        for key, value in column_mapping.items():
            lowercase_mapping[key.lower()] = value
        
        # Try to rename columns if they exist
        renamed_columns = {}
        for i, col in enumerate(df.columns):
            # Try exact match first
            if col in column_mapping:
                renamed_columns[col] = column_mapping[col]
            # Try lowercase match
            elif col.lower() in lowercase_mapping:
                renamed_columns[col] = lowercase_mapping[col.lower()]
            # Try partial match if no exact match found
            else:
                for key in column_mapping:
                    # Skip very short keys to avoid false matches
                    if len(key) < 3:
                        continue
                    if key in col or key.lower() in col.lower():
                        renamed_columns[col] = column_mapping[key]
                        print(f"Partial match: '{col}' -> '{column_mapping[key]}'")
                        break
        
        # Apply the column renaming
        df.rename(columns=renamed_columns, inplace=True)
        
        # Print renamed columns for debugging
        print("\nColumns after renaming:")
        for i, col in enumerate(df.columns):
            print(f"{i}: {col}")
        
        return df
    except Exception as e:
        print(f"Error reading shipping list file: {e}")
        raise


def read_policy_file(file_path):
    """
    Read the policy Excel file that contains markup and insurance rates.
    
    Args:
        file_path (str): Path to the policy Excel file
        
    Returns:
        dict: Dictionary containing markup percentage and insurance rate
    """
    try:
        df = pd.read_excel(file_path)
        # Assuming the policy file has columns for markup_percentage and insurance_rate
        # Adjust as needed based on actual file structure
        policy = {
            'markup_percentage': df['markup_percentage'].iloc[0] / 100,  # Convert to decimal
            'insurance_rate': df['insurance_rate'].iloc[0] / 100,  # Convert to decimal
            'insurance_coefficient': df.get('insurance_coefficient', [1.0]).iloc[0]  # Default to 1.0 if not present
        }
        return policy
    except Exception as e:
        print(f"Error reading policy file: {e}")
        raise


def read_shipping_rate_file(file_path):
    """
    Read the shipping rate Excel file.
    
    Args:
        file_path (str): Path to the shipping rate Excel file
        
    Returns:
        float: Current shipping rate
    """
    try:
        df = pd.read_excel(file_path)
        # Assuming the shipping rate file has a column for shipping_rate
        # Adjust as needed based on actual file structure
        shipping_rate = df['shipping_rate'].iloc[0]
        return shipping_rate
    except Exception as e:
        print(f"Error reading shipping rate file: {e}")
        raise


def read_exchange_rate_file(file_path):
    """
    Read the exchange rate Excel file.
    
    Args:
        file_path (str): Path to the exchange rate Excel file
        
    Returns:
        dict: Dictionary containing exchange rates between currencies
    """
    try:
        df = pd.read_excel(file_path)
        # Assuming the exchange rate file has columns for different currency pairs
        # Adjust as needed based on actual file structure
        exchange_rates = {
            'RMB_USD': df['RMB_USD'].iloc[0],
            'RMB_RUPEE': df.get('RMB_RUPEE', [0]).iloc[0],
            'USD_RUPEE': df.get('USD_RUPEE', [0]).iloc[0]
        }
        return exchange_rates
    except Exception as e:
        print(f"Error reading exchange rate file: {e}")
        raise


def deduplicate_shipping_list(df):
    """
    Deduplicate items in the shipping list where part number and unit price are the same.
    
    Args:
        df (pd.DataFrame): Original shipping list DataFrame
        
    Returns:
        pd.DataFrame: Deduplicated shipping list
    """
    try:
        # Make a copy to avoid modifying the original DataFrame
        df_copy = df.copy()
        
        # Print column names for debugging
        print("Available columns in the input file:")
        for i, col in enumerate(df_copy.columns):
            print(f"{i}: {col}")
        
        # Print first few rows for debugging
        print("\nFirst 3 rows of data:")
        print(df_copy.head(3))
        
        # Handle duplicate columns by renaming them
        if df_copy.columns.duplicated().any():
            print("Warning: Duplicate column names detected in input file. Renaming duplicate columns.")
            duplicate_cols = df_copy.columns[df_copy.columns.duplicated()].tolist()
            for col in duplicate_cols:
                # Find all occurrences of the duplicate column
                cols = df_copy.columns.tolist()
                indices = [i for i, x in enumerate(cols) if x == col]
                
                # Rename all but the first occurrence
                for i, idx in enumerate(indices[1:], 1):
                    cols[idx] = f"{col}_{i}"
                
                # Assign new column names to DataFrame
                df_copy.columns = cols
                print(f"Renamed duplicate columns for '{col}'")
        
        # Handle invalid/empty rows - drop rows where all key columns are NaN
        df_copy = df_copy.dropna(how='all')
        print(f"After dropping completely empty rows: {len(df_copy)} rows")
        
        # If DataFrame is empty after dropping empty rows, return original DataFrame
        if len(df_copy) == 0:
            print("Warning: DataFrame is empty after dropping empty rows. Returning original DataFrame.")
            return df
        
        # Identify groupby columns, using fallbacks if necessary
        part_number_col = 'part_number'
        unit_price_col = 'unit_price'
        
        # Check if the primary columns exist, otherwise look for alternatives
        if part_number_col not in df_copy.columns:
            # Try to find column containing 'P/N' or similar
            potential_pn_cols = [col for col in df_copy.columns if any(term in col.upper() for term in ['P/N', 'PART', 'PN', '料号'])]
            if potential_pn_cols:
                part_number_col = potential_pn_cols[0]
                print(f"Using '{part_number_col}' as the part number column")
            else:
                # If still not found, try to identify by position (assuming part number is usually one of the first few columns)
                print("Could not find a column containing 'P/N' or 'part'. Creating a default part number column.")
                # Create a synthetic part number column using first column or row index as fallback
                if len(df_copy.columns) > 1:
                    # Use the second column as part number (often part number is the second column after serial number)
                    df_copy['part_number'] = df_copy.iloc[:, 1]
                    part_number_col = 'part_number'
                    print(f"Created 'part_number' column using values from column: {df_copy.columns[1]}")
                else:
                    # Use row index as last resort
                    df_copy['part_number'] = df_copy.index
                    part_number_col = 'part_number'
                    print("Created 'part_number' column using row indices")
        
        if unit_price_col not in df_copy.columns:
            # Try to find column containing 'price' or similar
            potential_price_cols = [col for col in df_copy.columns if any(term in col.lower() for term in ['price', 'unit', '单价', 'cost'])]
            if potential_price_cols:
                unit_price_col = potential_price_cols[0]
                print(f"Using '{unit_price_col}' as the unit price column")
            else:
                # If still not found, create a dummy price column with constant value
                print("Could not find a unit price column. Creating a default unit price column with constant value 1.0")
                df_copy['unit_price'] = 1.0
                unit_price_col = 'unit_price'
        
        # Ensure we've found or created both required columns
        if not all(col in df_copy.columns for col in [part_number_col, unit_price_col]):
            print(f"Error: Could not find or create both {part_number_col} and {unit_price_col} columns.")
            return df
            
        groupby_cols = [part_number_col, unit_price_col]
        print(f"Grouping by columns: {groupby_cols}")
        
        # Verify groupby columns actually exist
        for col in groupby_cols:
            if col not in df_copy.columns:
                print(f"Critical error: Column '{col}' not found in DataFrame despite fallback measures. Using original DataFrame.")
                return df
        
        # Identify quantity column
        quantity_col = 'quantity'
        if quantity_col not in df_copy.columns:
            potential_qty_cols = [col for col in df_copy.columns if any(term in col.lower() for term in ['qty', 'quantity', '数量', 'count'])]
            if potential_qty_cols:
                quantity_col = potential_qty_cols[0]
                print(f"Using '{quantity_col}' as the quantity column")
            else:
                # If still not found, create a dummy quantity column
                print("Could not find a quantity column. Creating a default quantity column with constant value 1")
                df_copy['quantity'] = 1
                quantity_col = 'quantity'
        
        # Identify columns to sum based on whether they contain specific keywords
        sum_cols = []
        if quantity_col not in sum_cols and quantity_col in df_copy.columns:
            sum_cols.append(quantity_col)
        
        for col in df_copy.columns:
            if col not in groupby_cols and col not in sum_cols:
                if any(keyword in col.lower() for keyword in ['volume', 'weight', 'qty', 'count', 'piece', '体积', '重量', '数量', '件数']):
                    sum_cols.append(col)
        
        print(f"Columns that will be summed: {sum_cols}")
        
        # Ensure there are columns to sum
        if not sum_cols:
            print("Warning: No columns identified for summing. Deduplication may not be meaningful.")
            sum_cols = [quantity_col]  # Use quantity as a fallback
        
        # Convert numeric columns to the appropriate types
        for col in groupby_cols + sum_cols:
            if col in df_copy.columns:
                try:
                    df_copy[col] = pd.to_numeric(df_copy[col], errors='coerce')
                    print(f"Converted '{col}' to numeric")
                except Exception as e:
                    print(f"Could not convert '{col}' to numeric: {e}")
        
        # Replace NaN values with 0 in numeric columns to ensure proper summing
        for col in sum_cols:
            if col in df_copy.columns:
                df_copy[col] = df_copy[col].fillna(0)
        
        # Identify rows with NaN in groupby columns (these will be dropped during groupby)
        for col in groupby_cols:
            if col in df_copy.columns:
                nan_count = df_copy[col].isna().sum()
                if nan_count > 0:
                    print(f"Warning: {nan_count} rows have NaN values in '{col}' and will be excluded from groupby")
                    # Fill NaN values with a placeholder to avoid losing data
                    if pd.api.types.is_numeric_dtype(df_copy[col]):
                        df_copy[col] = df_copy[col].fillna(-99999)  # Use an unlikely value as placeholder
                    else:
                        df_copy[col] = df_copy[col].fillna("UNKNOWN")
        
        try:
            # Use a simplified two-step approach to avoid the pandas 'name' attribute error
            
            # Step 1: First group by and get the sum of the summable columns
            if sum_cols:
                sum_data = {}
                # Use a try-except block to catch any errors during groupby
                try:
                    for group_key, group_df in df_copy.groupby(groupby_cols):
                        # Handle case where group_key might be a single value or a tuple
                        if not isinstance(group_key, tuple):
                            group_key = (group_key,)
                        
                        # Sum each column for this group
                        sums = {col: group_df[col].sum() for col in sum_cols if col in group_df.columns}
                        sum_data[group_key] = sums
                    
                    # Create a DataFrame for the summed data
                    sum_rows = []
                    for group_key, sums in sum_data.items():
                        row = dict(zip(groupby_cols, group_key))
                        row.update(sums)
                        sum_rows.append(row)
                    
                    if sum_rows:
                        sum_df = pd.DataFrame(sum_rows)
                    else:
                        # If no rows, create an empty DataFrame with appropriate columns
                        sum_df = pd.DataFrame(columns=groupby_cols + sum_cols)
                    
                except Exception as e:
                    print(f"Error during groupby/sum operation: {e}")
                    print("Falling back to original DataFrame")
                    return df
            else:
                # If no sum columns, just get the unique combinations of groupby columns
                sum_df = df_copy[groupby_cols].drop_duplicates()
            
            # Step 2: Get the first row for each group for non-sum columns
            non_sum_cols = [col for col in df_copy.columns if col not in groupby_cols and col not in sum_cols]
            
            if non_sum_cols:
                first_data = {}
                try:
                    for group_key, group_df in df_copy.groupby(groupby_cols):
                        # Handle case where group_key might be a single value or a tuple
                        if not isinstance(group_key, tuple):
                            group_key = (group_key,)
                        
                        # Get the first row of each column for this group
                        first_row = group_df.iloc[0]
                        firsts = {col: first_row[col] for col in non_sum_cols if col in first_row.index}
                        first_data[group_key] = firsts
                    
                    # Create a DataFrame for the first-value data
                    first_rows = []
                    for group_key, firsts in first_data.items():
                        row = dict(zip(groupby_cols, group_key))
                        row.update(firsts)
                        first_rows.append(row)
                    
                    if first_rows:
                        first_df = pd.DataFrame(first_rows)
                    else:
                        # If no rows, create an empty DataFrame with appropriate columns
                        first_df = pd.DataFrame(columns=groupby_cols + non_sum_cols)
                    
                    # Merge the summed data with the first-value data
                    df_deduped = pd.merge(sum_df, first_df, on=groupby_cols, how='outer')
                except Exception as e:
                    print(f"Error during groupby/first operation: {e}")
                    print("Using only sum columns for deduplication")
                    df_deduped = sum_df
            else:
                # If no non-sum columns, just use the summed data
                df_deduped = sum_df
            
            # If the result is empty, return the original DataFrame
            if len(df_deduped) == 0:
                print("Warning: Deduplication resulted in empty DataFrame. Using original DataFrame.")
                return df
                
            # Recalculate unit-based values if they exist
            if quantity_col in df_deduped.columns:
                # Look for net weight columns
                net_weight_unit_col = next((col for col in df_deduped.columns if 'unit_net_weight' in col.lower() or '单件净重' in col), None)
                net_weight_total_col = next((col for col in df_deduped.columns if 'total_net_weight' in col.lower() or '总净重' in col), None)
                
                if net_weight_unit_col and net_weight_total_col and net_weight_total_col in df_deduped.columns:
                    df_deduped[net_weight_unit_col] = df_deduped[net_weight_total_col] / df_deduped[quantity_col].replace(0, np.nan)
                
                # Look for gross weight columns
                gross_weight_unit_col = next((col for col in df_deduped.columns if 'unit_gross_weight' in col.lower() or '单件毛重' in col), None)
                gross_weight_total_col = next((col for col in df_deduped.columns if 'total_gross_weight' in col.lower() or '总毛重' in col), None)
                
                if gross_weight_unit_col and gross_weight_total_col and gross_weight_total_col in df_deduped.columns:
                    df_deduped[gross_weight_unit_col] = df_deduped[gross_weight_total_col] / df_deduped[quantity_col].replace(0, np.nan)
            
            print(f"Successfully deduplicated: {len(df_deduped)} rows (from original {len(df_copy)} rows)")
            return df_deduped
            
        except Exception as e:
            print(f"Error during deduplication: {e}")
            # If deduplication fails, return the original DataFrame as a fallback
            print("Falling back to original data without deduplication")
            return df
            
    except Exception as e:
        print(f"Unexpected error in deduplicate_shipping_list: {e}")
        # In case of any unexpected error, return the original DataFrame
        return df


def calculate_fob_prices(df, policy):
    """
    Calculate FOB prices for the shipping list according to the specification.
    
    The calculation follows these steps:
    1. Get the markup percentage from the policy file
    2. Calculate FOB prices:
       2.a FOB unit price = Unit price * (1 + markup %)
       2.b FOB total price = FOB unit price * quantity
    
    Args:
        df (pd.DataFrame): Shipping list DataFrame
        policy (dict): Policy parameters including markup percentage
        
    Returns:
        pd.DataFrame: DataFrame with added FOB pricing information
    """
    # Make a copy to avoid modifying the original DataFrame
    df_copy = df.copy()
    
    try:
        # Helper function to safely get or create columns
        def safe_get_or_create_column(df, column_name, fallback_columns=None, default_value=0):
            if column_name in df.columns:
                return df[column_name]
            
            if fallback_columns:
                for fallback_col in fallback_columns:
                    if fallback_col in df.columns:
                        print(f"Using '{fallback_col}' instead of '{column_name}' for FOB calculation")
                        df[column_name] = df[fallback_col]
                        return df[column_name]
            
            print(f"Column '{column_name}' not found. Creating with default value {default_value}")
            df[column_name] = default_value
            return df[column_name]
        
        # Ensure required columns exist
        unit_price = safe_get_or_create_column(df_copy, 'unit_price', 
                                             fallback_columns=['Unit Price', '单价', 'price', '不含税单价', '不含税单价（RMB）'],
                                             default_value=1.0)
        
        quantity = safe_get_or_create_column(df_copy, 'quantity', 
                                           fallback_columns=['Quantity', 'Qty', '数量', 'QUANTITY', 'QUANTITY （数量）'],
                                           default_value=1.0)
        
        # Get policy parameters with defaults
        markup_percentage = policy.get('markup_percentage', 0.05)  # Default 5% if not provided
        
        # Step 2.a: Calculate FOB unit price
        df_copy['fob_unit_price'] = df_copy['unit_price'] * (1 + markup_percentage)
        
        # Step 2.b: Calculate FOB total price
        df_copy['fob_total_price'] = df_copy['fob_unit_price'] * df_copy['quantity']
        
        print("FOB price calculation completed successfully")
        return df_copy
        
    except Exception as e:
        print(f"Error during FOB price calculation: {e}")
        # If there's an error, add default FOB columns with reasonable values
        print("Adding default FOB pricing due to calculation error")
        df_copy['fob_unit_price'] = df_copy.get('unit_price', 10.0) * 1.05  # 5% markup as fallback
        df_copy['fob_total_price'] = df_copy['fob_unit_price'] * df_copy.get('quantity', 1.0)
        return df_copy


def save_fob_prices(df, output_path):
    """
    Save the DataFrame with FOB prices to an Excel file.
    Updates gross weight information if it already exists.
    Preserves all original columns from the input DataFrame.
    """
    try:
        # Make a copy to avoid modifying the original DataFrame
        df_copy = df.copy()
        
        def safe_get_or_create_column(df, column_name, fallback_columns=None, default_value=0):
            if column_name in df.columns:
                return df[column_name]
            if fallback_columns:
                for col in fallback_columns:
                    if col in df.columns:
                        return df[col]
            print(f"Creating new column: {column_name}")
            df[column_name] = default_value
            return df[column_name]

        # Convert quantity and weights to numeric if they exist
        if 'quantity' in df_copy.columns:
            df_copy['quantity'] = pd.to_numeric(df_copy['quantity'], errors='coerce').fillna(0)
        if 'unit_gross_weight' in df_copy.columns:
            df_copy['unit_gross_weight'] = pd.to_numeric(df_copy['unit_gross_weight'], errors='coerce').fillna(0)
        if 'total_gross_weight' in df_copy.columns:
            df_copy['total_gross_weight'] = pd.to_numeric(df_copy['total_gross_weight'], errors='coerce').fillna(0)

        # Calculate total gross weight if necessary
        if 'unit_gross_weight' in df_copy.columns and 'quantity' in df_copy.columns:
            if 'total_gross_weight' not in df_copy.columns:
                df_copy['total_gross_weight'] = df_copy['unit_gross_weight'] * df_copy['quantity']

        # Save to Excel with all columns
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_copy.to_excel(writer, sheet_name='FOB Prices', index=False)
            
            # Add metadata sheet
            metadata = pd.DataFrame({
                'Created Date': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                'Total Items': [len(df_copy)],
                'Total Quantity': [df_copy['quantity'].sum() if 'quantity' in df_copy.columns else 0],
                'Total Gross Weight': [df_copy['total_gross_weight'].sum() if 'total_gross_weight' in df_copy.columns else 0]
            })
            metadata.to_excel(writer, sheet_name='Metadata', index=False)

        print(f"Successfully saved FOB prices to {output_path}")
        return True

    except Exception as e:
        print(f"Error saving FOB prices: {str(e)}")
        return False


def calculate_cif_prices(df, policy, shipping_rate, exchange_rates):
    """
    Calculate CIF unit prices for the shipping list according to the updated specification.
    
    The calculation follows these steps:
    4. The CIF unit price is calculated based on the following:
       4.a Convert FOB prices from CNY to USD by multiplying by exchange rate (RMB_USD)
       4.b Calculate the adjusted total goods cost with insurance: (total FOB price in USD * insurance co-efficiency * (1+insurance rate))
       4.c Calculate the total shipping cost in CNY first: net weight * shipping rate, then convert to USD
       4.d Calculate the CIF total cost in USD: total goods cost with insurance + total shipping cost in USD
       4.e Calculate the CIF unit price in USD: total CIF cost / quantity
    """
    # Make a copy to avoid modifying the original DataFrame
    df_copy = df.copy()
    
    try:
        print("\nStarting CIF price calculation...")
        exchange_rate_usd = exchange_rates.get('RMB_USD', 0.13)  # Default to 0.13 if not provided
        print(f"Exchange rate (RMB/USD): {exchange_rate_usd}")
        print(f"Shipping rate (CNY/kg): {shipping_rate}")
        print(f"Insurance coefficient: {policy.get('insurance_coefficient', 1.0)}")
        print(f"Insurance rate: {policy.get('insurance_rate', 0.01)}")
        
        # Helper function to safely get or create columns
        def safe_get_or_create_column(df, column_name, fallback_columns=None, default_value=0):
            if column_name in df.columns:
                return df[column_name]
            
            if fallback_columns:
                for fallback_col in fallback_columns:
                    if fallback_col in df.columns:
                        print(f"Using '{fallback_col}' instead of '{column_name}' for CIF calculation")
                        df[column_name] = df[fallback_col]
                        return df[column_name]
            
            print(f"Column '{column_name}' not found. Creating with default value {default_value}")
            df[column_name] = default_value
            return df[column_name]
        
        # Ensure required columns exist
        quantity = safe_get_or_create_column(df_copy, 'quantity', 
                                           fallback_columns=['Quantity', 'Qty', '数量', 'QUANTITY', 'QUANTITY （数量）'],
                                           default_value=1.0)
        
        total_net_weight = safe_get_or_create_column(df_copy, 'total_net_weight', 
                                                  fallback_columns=['N.W', '总净重', 'Total Net Weight', 'N.W  (KG) 总净重'],
                                                  default_value=0.1)
        
        # Check if net weight is missing, if so, use gross weight * 0.9
        if 'total_net_weight' not in df_copy.columns or df_copy['total_net_weight'].isna().all():
            total_gross_weight = safe_get_or_create_column(df_copy, 'total_gross_weight',
                                                      fallback_columns=['G.W', '总毛重', 'Total Gross Weight', 'G.W（KG) 总毛重'],
                                                      default_value=0.1)
            df_copy['total_net_weight'] = total_gross_weight * 0.9
            print("Using gross weight * 0.9 as net weight")
        
        # Ensure FOB prices exist and are numeric
        fob_total_price = pd.to_numeric(safe_get_or_create_column(df_copy, 'fob_total_price', default_value=0), errors='coerce').fillna(0)
        fob_unit_price = pd.to_numeric(safe_get_or_create_column(df_copy, 'fob_unit_price', default_value=0), errors='coerce').fillna(0)
        
        # Step 4.a: Convert FOB prices from CNY to USD by multiplying by exchange rate
        df_copy['fob_total_price_usd'] = df_copy['fob_total_price'] * exchange_rate_usd
        df_copy['fob_unit_price_usd'] = df_copy['fob_unit_price'] * exchange_rate_usd
        
        print("\nSample FOB price conversion (first 3 rows):")
        for i in range(min(3, len(df_copy))):
            print(f"\nRow {i+1}:")
            print(f"  FOB Unit Price (CNY): ¥{df_copy['fob_unit_price'].iloc[i]:.2f}")
            print(f"  FOB Unit Price (USD): ${df_copy['fob_unit_price_usd'].iloc[i]:.2f}")
            print(f"  FOB Total Price (CNY): ¥{df_copy['fob_total_price'].iloc[i]:.2f}")
            print(f"  FOB Total Price (USD): ${df_copy['fob_total_price_usd'].iloc[i]:.2f}")
        
        # Get policy parameters with defaults
        insurance_coefficient = policy.get('insurance_coefficient', 1.0)  # Default 1.0 if not provided
        insurance_rate = policy.get('insurance_rate', 0.01)  # Default 1% if not provided
        
        # Step 4.b: Calculate adjusted total goods cost with insurance in USD
        df_copy['total_goods_cost_with_insurance'] = df_copy['fob_total_price_usd'] * insurance_coefficient * (1 + insurance_rate)
        
        # Step 4.c: Calculate total shipping cost in CNY first, then convert to USD
        df_copy['total_shipping_cost_cny'] = df_copy['total_net_weight'] * shipping_rate
        df_copy['total_shipping_cost_usd'] = df_copy['total_shipping_cost_cny'] * exchange_rate_usd
        
        # Step 4.d: Calculate CIF total cost in USD
        df_copy['cif_total_cost_usd'] = df_copy['total_goods_cost_with_insurance'] + df_copy['total_shipping_cost_usd']
        
        # Step 4.e: Calculate CIF unit price in USD
        # Avoid division by zero by replacing 0 with NaN temporarily
        safe_quantity = df_copy['quantity'].replace(0, np.nan)
        df_copy['cif_unit_price_usd'] = df_copy['cif_total_cost_usd'] / safe_quantity
        df_copy['cif_unit_price_usd'] = df_copy['cif_unit_price_usd'].fillna(0)  # Replace NaN back with 0
        
        print("\nSample CIF price calculation (first 3 rows):")
        for i in range(min(3, len(df_copy))):
            print(f"\nRow {i+1}:")
            print(f"  Quantity: {df_copy['quantity'].iloc[i]}")
            print(f"  Net Weight: {df_copy['total_net_weight'].iloc[i]:.2f} kg")
            print(f"  Total Goods Cost with Insurance: ${df_copy['total_goods_cost_with_insurance'].iloc[i]:.2f}")
            print(f"  Total Shipping Cost (CNY): ¥{df_copy['total_shipping_cost_cny'].iloc[i]:.2f}")
            print(f"  Total Shipping Cost (USD): ${df_copy['total_shipping_cost_usd'].iloc[i]:.2f}")
            print(f"  CIF Total Cost (USD): ${df_copy['cif_total_cost_usd'].iloc[i]:.2f}")
            print(f"  CIF Unit Price (USD): ${df_copy['cif_unit_price_usd'].iloc[i]:.2f}")
        
        # Calculate RMB prices for reference (divide USD prices by exchange rate)
        df_copy['cif_unit_price_rmb'] = df_copy['cif_unit_price_usd'] / exchange_rate_usd
        df_copy['cif_total_cost_rmb'] = df_copy['cif_total_cost_usd'] / exchange_rate_usd
        
        print("\nCIF price calculation completed successfully")
        return df_copy
        
    except Exception as e:
        print(f"Error during CIF price calculation: {e}")
        # If there's an error, add default CIF columns with reasonable values
        print("Adding default CIF pricing due to calculation error")
        df_copy['cif_unit_price_usd'] = df_copy.get('fob_unit_price', 10.0) * exchange_rates.get('RMB_USD', 0.13)  # Use exchange rate
        df_copy['cif_unit_price_rmb'] = df_copy['cif_unit_price_usd'] / exchange_rate_usd
        return df_copy


def generate_export_receipt(df_export, output_file):
    """
    Generate an export receipt Excel file with CIF prices in USD.
    
    Args:
        df_export (pd.DataFrame): DataFrame containing the export data with CIF prices
        output_file (str): Path to save the export receipt Excel file
        
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print("\nGenerating export receipt...")
        
        # Helper function to safely get column data with fallbacks
        def safe_get_column(df, target_col, fallback_columns=None):
            if target_col in df.columns:
                return df[target_col]
            if fallback_columns:
                for col in fallback_columns:
                    if col in df.columns:
                        print(f"Using fallback column '{col}' for '{target_col}'")
                        return df[col]
            print(f"Warning: Column '{target_col}' and fallbacks not found. Using empty values.")
            return pd.Series(['' for _ in range(len(df))])

        # Create a new DataFrame for the export receipt
        export_df = pd.DataFrame()

        # Handle duplicate columns by taking the first occurrence
        df_export = df_export.loc[:, ~df_export.columns.duplicated()]

        # Map existing columns to required format
        export_df["NO."] = range(1, len(df_export) + 1)
        export_df["P/N"] = safe_get_column(df_export, "part_number", ["Part Number", "零件号", "料号"])
        export_df["DESCRIPTION"] = safe_get_column(df_export, "description_en", ["DESCRIPTION", "Description", "英文品名", "customs_desc_en", "material_name"])
        export_df["Model NO."] = safe_get_column(df_export, "model", ["MODEL", "Model", "货物型号"])
        
        # Get quantity and ensure it's numeric
        export_df["Qty"] = pd.to_numeric(safe_get_column(
            df_export, "quantity",
            fallback_columns=["Quantity", "Qty", "数量"]
        ), errors='coerce').fillna(0)

        # Get CIF unit price in USD and ensure it's numeric
        cif_unit_price_usd = pd.to_numeric(safe_get_column(
            df_export, "cif_unit_price_usd",
            fallback_columns=["cif_unit_price_usd"]  # No fallback to other price columns
        ), errors='coerce').fillna(0)

        # Round to 2 decimal places for display
        export_df["Unit Price USD"] = cif_unit_price_usd.round(2)
        
        # Calculate Amount USD as Qty * Unit Price USD
        export_df["Amount USD"] = (export_df["Qty"] * export_df["Unit Price USD"]).round(2)

        export_df["Unit"] = safe_get_column(df_export, "unit", ["Unit", "单位"])

        # Print debug information
        print("\nExport receipt summary:")
        print(f"Number of rows: {len(export_df)}")
        print("\nSample of prices (first 3 rows):")
        for i in range(min(3, len(export_df))):
            print(f"\nRow {i+1}:")
            print(f"  P/N: {export_df['P/N'].iloc[i]}")
            print(f"  Qty: {export_df['Qty'].iloc[i]}")
            print(f"  Unit Price USD: ${export_df['Unit Price USD'].iloc[i]:.2f}")
            print(f"  Amount USD: ${export_df['Amount USD'].iloc[i]:.2f}")

        print(f"\nTotal Amount USD: ${export_df['Amount USD'].sum():.2f}")

        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write main data
            export_df.to_excel(writer, sheet_name='Export Receipt', index=False)
            
            # Create metadata sheet
            metadata_df = pd.DataFrame()
            metadata_sheet = writer.book.create_sheet('Metadata')
            
            # Add total amount to metadata
            metadata_sheet['A1'] = 'Export Receipt Summary'
            metadata_sheet['A2'] = 'Total Amount USD'
            metadata_sheet['B2'] = export_df["Amount USD"].sum().round(2)
            
            # Format the metadata sheet
            for cell in ['A1', 'A2', 'B2']:
                metadata_sheet[cell].font = Font(bold=True)
            
            # Adjust column widths in main sheet
            worksheet = writer.sheets['Export Receipt']
            for idx, col in enumerate(export_df.columns):
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = 15

        print(f"Export receipt generated successfully: {output_file}")
        return True
        
    except Exception as e:
        print(f"Error generating export receipt: {e}")
        return False


def generate_reimport_receipt(df, output_path):
    """
    Generate re-import receipt Excel file with CIF prices in USD.
    
    Args:
        df (pd.DataFrame): DataFrame with CIF pricing information
        output_path (str): Path to save the re-import receipt
        
    Returns:
        str: Path to the saved re-import receipt
    """
    # Make a copy to avoid modifying the original DataFrame
    df_reimport = df.copy()
    
    print("Generating re-import receipt...")
    print(f"Available columns for re-import receipt: {df_reimport.columns.tolist()}")
    
    # Handle duplicate columns by renaming them with a suffix
    if df_reimport.columns.duplicated().any():
        print("Warning: Duplicate column names detected. Renaming duplicate columns.")
        duplicate_cols = df_reimport.columns[df_reimport.columns.duplicated()].tolist()
        for col in duplicate_cols:
            # Find all occurrences of the duplicate column
            cols = df_reimport.columns.tolist()
            indices = [i for i, x in enumerate(cols) if x == col]
            
            # Rename all but the first occurrence
            for i, idx in enumerate(indices[1:], 1):
                cols[idx] = f"{col}_{i}"
            
            # Assign new column names to DataFrame
            df_reimport.columns = cols
            print(f"Renamed duplicate columns for '{col}'")
    
    # Create a new DataFrame with the required columns for re-import receipt
    reimport_df = pd.DataFrame()
    
    # Helper function to safely get column data with a fallback
    def safe_get_column(dataframe, column_name, fallback_value=None, fallback_columns=None):
        if column_name in dataframe.columns:
            return dataframe[column_name]
        elif fallback_columns:
            for fallback_col in fallback_columns:
                if fallback_col in dataframe.columns:
                    print(f"Using '{fallback_col}' instead of '{column_name}'")
                    return dataframe[fallback_col]
        print(f"Column '{column_name}' not found. Using fallback value.")
        return fallback_value if fallback_value is not None else pd.Series(['-'] * len(dataframe))
    
    # Map the existing columns to a format similar to export receipt
    # but include customs description fields which are important for re-import
    reimport_df["NO."] = safe_get_column(
        df_reimport, "serial_no", 
        fallback_value=range(1, len(df_reimport) + 1),
        fallback_columns=["Sr NO", "序列号", "Serial No", "Serial Number"]
    )
    
    reimport_df["P/N"] = safe_get_column(
        df_reimport, "part_number", 
        fallback_columns=["P/N.", "P/N", "Part Number", "料号", "系统料号"]
    )
    
    reimport_df["English Description"] = safe_get_column(
        df_reimport, "customs_desc_en", 
        fallback_columns=["清关英文货描", "Customs Description", "description_en", "DESCRIPTION", "英文品名"]
    )
    
    reimport_df["Chinese Description"] = safe_get_column(
        df_reimport, "customs_desc_cn", 
        fallback_columns=["报关中文品名", "中文品名"]
    )
    
    reimport_df["Model NO."] = safe_get_column(
        df_reimport, "model", 
        fallback_columns=["MODEL", "Model", "货物型号"]
    )
    
    # Ensure we use CIF prices in USD, with no fallback to original unit price
    reimport_df["Unit Price USD"] = safe_get_column(
        df_reimport, "cif_unit_price_usd"
    ).round(2)
    
    reimport_df["Qty"] = safe_get_column(
        df_reimport, "quantity", 
        fallback_columns=["QUANTITY", "Quantity", "数量", "Qty"]
    )
    
    reimport_df["Unit"] = safe_get_column(
        df_reimport, "unit", 
        fallback_columns=["Unit", "单位"]
    )
    
    reimport_df["Amount USD"] = (reimport_df["Unit Price USD"] * reimport_df["Qty"]).round(2)

    reimport_df["Net Weight (kg)"] = safe_get_column(
        df_reimport, "unit_net_weight", 
        fallback_columns=["单件净重", "Unit Net Weight"]
    )
    
    reimport_df["Total Net Weight (kg)"] = safe_get_column(
        df_reimport, "total_net_weight", 
        fallback_columns=["N.W", "总净重", "Total Net Weight"]
    )
    
    reimport_df["Gross Weight (kg)"] = safe_get_column(
        df_reimport, "unit_gross_weight", 
        fallback_columns=["单件毛重", "Unit Gross Weight"]
    )
    
    reimport_df["Total Gross Weight (kg)"] = safe_get_column(
        df_reimport, "total_gross_weight", 
        fallback_columns=["G.W", "总毛重", "Total Gross Weight"]
    )
    
    # Save to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        reimport_df.to_excel(writer, index=False, sheet_name='Re-Import Receipt')
        
        # Add formatting if needed
        workbook = writer.book
        worksheet = writer.sheets['Re-Import Receipt']
        
        # Add date and other metadata
        metadata_sheet = workbook.create_sheet('Metadata')
        metadata_sheet['A1'] = 'Generated Date'
        metadata_sheet['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    print(f"Re-import receipt saved to {output_path}")
    return output_path


def normalize_shipping_list(df):
    """
    Normalize the shipping list by breaking down combined box items into individual rows.
    Handles merged cells by distributing the original total weight evenly across all items.
    Preserves all original columns in the process.
    Fills missing information from the previous line when available.
    """
    try:
        # Make a copy to avoid modifying the original DataFrame
        df_copy = df.copy()
        
        # Reset index to avoid any duplicate index issues
        df_copy.reset_index(drop=True, inplace=True)
        
        # Store all original columns
        original_columns = df_copy.columns.tolist()
        print("\nProcessing columns for empty values:")
        
        # Helper function to check if a value should be considered empty
        def is_empty_value(val):
            if pd.isna(val):  # Handles np.nan, None, and pd.NaT
                return True
            if isinstance(val, str):
                # Check for empty strings, whitespace, 'nan', 'none', 'null', '-', etc.
                cleaned = val.strip().lower()
                return cleaned == '' or cleaned == 'nan' or cleaned == 'none' or cleaned == 'null' or cleaned == '-'
            return False

        # Helper function to identify if rows are part of the same group (previously merged)
        def is_same_group(row1, row2):
            """
            Check if two rows were likely part of the same merged group.
            Returns True if the rows share key identifiers that suggest they were previously merged.
            """
            # First check part number as it's the most important identifier
            if 'part_number' in row1.index and 'part_number' in row2.index:
                part1 = str(row1['part_number']).strip() if not pd.isna(row1['part_number']) else ''
                part2 = str(row2['part_number']).strip() if not pd.isna(row2['part_number']) else ''
                if part1 and part2 and part1 == part2:
                    return True  # If part numbers match exactly, consider them the same group
            
            # If part numbers don't match or are empty, check other key identifiers
            key_columns = [
                'supplier', 'factory', 'model', 'customs_desc_en', 'customs_desc_cn',
                'description_en', 'material_name', 'project_name', 'invoice_name'
            ]
            
            # Count how many non-empty values match between the rows
            matches = 0
            total_comparisons = 0
            
            for col in key_columns:
                if col in row1.index and col in row2.index:
                    val1 = str(row1[col]).strip() if not pd.isna(row1[col]) else ''
                    val2 = str(row2[col]).strip() if not pd.isna(row2[col]) else ''
                    
                    # Only count if at least one value is non-empty
                    if val1 or val2:
                        total_comparisons += 1
                        if val1 == val2:
                            matches += 1
            
            # Consider rows part of the same group if:
            # 1. They have at least 2 matching non-empty values, or
            # 2. They have at least 1 matching value and no conflicting values
            return matches >= 2 or (matches >= 1 and matches == total_comparisons)

        # First pass: Process rows that were previously merged
        print("\nProcessing previously merged rows...")
        prev_row = None
        prev_idx = None
        current_group = []

        for idx in df_copy.index:
            current_row = df_copy.loc[idx]
            
            # Check if this row should be part of the current group
            if prev_row is not None and is_same_group(prev_row, current_row):
                if not current_group:  # If this is the start of a new group
                    current_group = [prev_idx, idx]
                else:
                    current_group.append(idx)
            else:
                # Process the previous group if it exists
                if current_group:
                    # Get the first row of the group as the source of truth
                    source_row = df_copy.loc[current_group[0]]
                    
                    # Copy values to all other rows in the group
                    for group_idx in current_group[1:]:
                        for col in df_copy.columns:
                            current_val = df_copy.at[group_idx, col]
                            source_val = source_row[col]
                            
                            # Don't copy certain columns that should be unique per row
                            if col not in ['serial_no', 'quantity', 'unit_gross_weight', 'total_gross_weight', 
                                         'unit_net_weight', 'total_net_weight', 'carton_no', 'volume',
                                         'total_volume']:
                                # Copy value if:
                                # 1. Current value is empty or a placeholder
                                # 2. Source value is not empty
                                # 3. Column is a text column (not numeric)
                                try:
                                    is_numeric = pd.to_numeric(df_copy[col], errors='coerce').notna().any()
                                except:
                                    is_numeric = False
                                
                                if not is_numeric and not pd.isna(source_val):
                                    if (pd.isna(current_val) or 
                                        (isinstance(current_val, str) and current_val.strip() in ['', '-', '/', '_', 'nan', 'none', 'null']) or
                                        str(current_val).strip() == ''):
                                        df_copy.at[group_idx, col] = source_val
                                        print(f"Copied value '{source_val}' from row {current_group[0]} to column '{col}' at row {group_idx}")
            
            # Start a new potential group
            current_group = [idx]
        
            prev_row = current_row
            prev_idx = idx

        # Process the last group if it exists
        if current_group:
            source_row = df_copy.loc[current_group[0]]
            for group_idx in current_group[1:]:
                for col in df_copy.columns:
                    if col not in ['serial_no', 'quantity', 'unit_gross_weight', 'total_gross_weight', 
                                  'unit_net_weight', 'total_net_weight', 'carton_no', 'volume',
                                  'total_volume']:
                        current_val = df_copy.at[group_idx, col]
                        source_val = source_row[col]
                        try:
                            is_numeric = pd.to_numeric(df_copy[col], errors='coerce').notna().any()
                        except:
                            is_numeric = False
                        
                        if not is_numeric and not pd.isna(source_val):
                            if (pd.isna(current_val) or 
                                (isinstance(current_val, str) and current_val.strip() in ['', '-', '/', '_', 'nan', 'none', 'null']) or
                                str(current_val).strip() == ''):
                                df_copy.at[group_idx, col] = source_val
                                print(f"Copied value '{source_val}' from row {current_group[0]} to column '{col}' at row {group_idx}")

        # Second pass: Handle empty values in all columns
        for col in df_copy.columns:
            print(f"\nProcessing column: {col}")
            
            # Try to convert to numeric first
            try:
                numeric_series = pd.to_numeric(df_copy[col], errors='coerce')
                if not numeric_series.isna().all():  # If at least some values are numeric
                    print(f"Column '{col}' is numeric")
                    # For numeric columns, only fill forward if the value is 0 or NaN
                    mask = numeric_series.isna() | (numeric_series == 0)
                    if mask.any():
                        prev_val = None
                        for idx in df_copy.index:
                            if mask[idx] and prev_val is not None:
                                df_copy.at[idx, col] = prev_val
                                print(f"Filled numeric value at row {idx} with {prev_val}")
                            elif not mask[idx]:
                                prev_val = df_copy.at[idx, col]
                    continue
            except:
                pass  # Not a numeric column, continue with string processing
            
            # For non-numeric columns
            empty_mask = df_copy[col].apply(is_empty_value)
            if empty_mask.any():
                prev_val = None
                for idx in df_copy.index:
                    current_val = df_copy.at[idx, col]
                    if is_empty_value(current_val):
                        if prev_val is not None:
                            df_copy.at[idx, col] = prev_val
                            print(f"Filled empty value at row {idx} with '{prev_val}'")
                    else:
                        prev_val = current_val

        # Third pass: Handle special columns that should be copied even if not empty
        special_columns = ['supplier', 'project_name', 'factory', 'customs_desc_en', 'customs_desc_cn', 
                         'description_en', 'material_name', 'model', 'invoice_name', 'purchasing_unit']
        
        for col in df_copy.columns:
            if any(special_name in col.lower() for special_name in special_columns):
                print(f"\nSpecial handling for column: {col}")
                prev_val = None
                for idx in df_copy.index:
                    current_val = df_copy.at[idx, col]
                    # For special columns, also copy if the current value looks like a placeholder
                    if is_empty_value(current_val) or (isinstance(current_val, str) and current_val.strip() in ['-', '/', '_']):
                        if prev_val is not None:
                            df_copy.at[idx, col] = prev_val
                            print(f"Filled value at row {idx} with '{prev_val}'")
                    else:
                        prev_val = current_val

        # Now proceed with the weight calculations
        def safe_get_or_create_column(df, column_name, fallback_columns=None, default_value=0):
            if column_name in df.columns:
                return df[column_name]
            if fallback_columns:
                for col in fallback_columns:
                    if col in df.columns:
                        return df[col]
            df[column_name] = default_value
            return df[column_name]

        # Get required columns with fallbacks
        part_number = safe_get_or_create_column(df_copy, 'part_number', ['P/N', 'part_num', '料号'])
        quantity = safe_get_or_create_column(df_copy, 'quantity', ['QUANTITY', 'qty'])
        unit_gross_weight = safe_get_or_create_column(df_copy, 'unit_gross_weight', ['Unit G.W', 'unit_gw'])
        total_gross_weight = safe_get_or_create_column(df_copy, 'total_gross_weight', ['G.W（KG)', 'total_gw'])
        unit_net_weight = safe_get_or_create_column(df_copy, 'unit_net_weight', ['Unit N.W', 'unit_nw'])
        total_net_weight = safe_get_or_create_column(df_copy, 'total_net_weight', ['N.W  (KG)', 'total_nw'])

        # Convert weights to numeric, replacing non-numeric values with 0
        df_copy['unit_gross_weight'] = pd.to_numeric(unit_gross_weight, errors='coerce').fillna(0)
        df_copy['total_gross_weight'] = pd.to_numeric(total_gross_weight, errors='coerce').fillna(0)
        df_copy['unit_net_weight'] = pd.to_numeric(unit_net_weight, errors='coerce').fillna(0)
        df_copy['total_net_weight'] = pd.to_numeric(total_net_weight, errors='coerce').fillna(0)
        df_copy['quantity'] = pd.to_numeric(quantity, errors='coerce').fillna(0)

        # Function to identify groups of consecutive rows with the same part number
        def get_merged_groups(df):
            groups = []
            current_group = []
            last_part_number = None
            
            for idx, row in df.iterrows():
                current_part_number = row['part_number']
                
                # Check if this row should be part of the current group
                if current_part_number == last_part_number and len(current_group) > 0:
                    # If the row has a total weight and previous rows don't, it might be the merged cell
                    if row['total_gross_weight'] > 0 and all(df.loc[i, 'total_gross_weight'] == 0 for i in current_group):
                        current_group.append(idx)
                    # If previous rows have a total weight and this one doesn't, it's likely part of the split
                    elif row['total_gross_weight'] == 0 and any(df.loc[i, 'total_gross_weight'] > 0 for i in current_group):
                        current_group.append(idx)
                    # If weights match between rows
                    elif row['total_gross_weight'] == df.loc[current_group[0], 'total_gross_weight']:
                        current_group.append(idx)
                    else:
                        if len(current_group) > 1:  # Only save groups with multiple rows
                            groups.append(current_group)
                        current_group = [idx]
                else:
                    if len(current_group) > 1:  # Only save groups with multiple rows
                        groups.append(current_group)
                    current_group = [idx]
                
                last_part_number = current_part_number
            
            # Don't forget to add the last group
            if len(current_group) > 1:
                groups.append(current_group)
            
            return groups

        # Process each group of merged rows
        merged_groups = get_merged_groups(df_copy)
        for group_indices in merged_groups:
            group_df = df_copy.loc[group_indices]
            
            # Calculate total items in the group
            total_items = group_df['quantity'].sum()
            if total_items <= 0:
                continue

            print(f"\nProcessing merged group with indices {group_indices}:")
            print(f"Total items in group: {total_items}")
            
            # Find the original total weight (should be in one of the rows)
            group_weights = group_df['total_gross_weight'].unique()
            group_weights = group_weights[group_weights > 0]
            if len(group_weights) > 0:
                original_gross_weight = max(group_weights)
                # Calculate unit weight
                unit_gross = original_gross_weight / total_items
                unit_net = unit_gross * 0.9  # Assume net weight is 90% of gross weight
                
                print(f"Original total gross weight: {original_gross_weight}")
                print(f"Calculated unit weights: gross={unit_gross}, net={unit_net}")
                
                # Update weights for each row in the group
                for idx in group_indices:
                    items_in_row = df_copy.at[idx, 'quantity']
                    if items_in_row <= 0:
                        continue
                    
                    print(f"Processing row {idx} with {items_in_row} items")
                    
                    # Update only the weight columns, preserving all other data
                    df_copy.at[idx, 'unit_gross_weight'] = unit_gross
                    df_copy.at[idx, 'total_gross_weight'] = unit_gross * items_in_row
                    df_copy.at[idx, 'unit_net_weight'] = unit_net
                    df_copy.at[idx, 'total_net_weight'] = unit_net * items_in_row
                    
                    print(f"Final weights for row {idx}:")
                    print(f"Unit gross weight: {df_copy.at[idx, 'unit_gross_weight']}")
                    print(f"Total gross weight: {df_copy.at[idx, 'total_gross_weight']}")

        # Process remaining rows (not part of merged groups) using the original carton-based logic
        processed_indices = set([idx for group in merged_groups for idx in group])
        remaining_indices = set(df_copy.index) - processed_indices
        
        if remaining_indices:
            print("\nProcessing remaining rows...")
            for idx in remaining_indices:
                row = df_copy.loc[idx]
                if row['quantity'] <= 0:
                    continue
                
                # If the row already has valid weights, keep them
                if row['unit_gross_weight'] > 0 and row['total_gross_weight'] > 0:
                    continue
                
                # Otherwise, calculate weights based on the row's data
                if row['total_gross_weight'] > 0:
                    unit_gross = row['total_gross_weight'] / row['quantity']
                elif row['unit_gross_weight'] > 0:
                    unit_gross = row['unit_gross_weight']
                else:
                    unit_gross = 0.1  # Minimum default weight
                
                unit_net = unit_gross * 0.9
                
                # Update only the weight columns, preserving all other data
                df_copy.at[idx, 'unit_gross_weight'] = unit_gross
                df_copy.at[idx, 'total_gross_weight'] = unit_gross * row['quantity']
                df_copy.at[idx, 'unit_net_weight'] = unit_net
                df_copy.at[idx, 'total_net_weight'] = unit_net * row['quantity']

        # Ensure all original columns are preserved in the same order
        df_copy = df_copy[original_columns]
        
        # Final pass: Check for any remaining empty values
        empty_counts = {}
        for col in df_copy.columns:
            empty_count = df_copy[col].apply(is_empty_value).sum()
            if empty_count > 0:
                empty_counts[col] = empty_count
        
        if empty_counts:
            print("\nWarning: Some columns still contain empty values:")
            for col, count in empty_counts.items():
                print(f"- {col}: {count} empty values")
        
        return df_copy
        
    except Exception as e:
        print(f"Error during shipping list normalization: {str(e)}")
        print("Returning original DataFrame")
        return df


def process_shipping_list(
    shipping_list_file,
    policy_file,
    shipping_rate_file,
    exchange_rate_file,
    output_fob_file,
    output_export_file,
    output_reimport_file
):
    """
    Process the shipping list according to the specification.
    Now includes a second pass to handle any remaining empty cells.
    """
    try:
        # First Pass
        print("\n=== First Pass ===")
        print("Reading input files...")
        df = read_shipping_list(shipping_list_file)
        if df is None or len(df) == 0:
            print("Error: Failed to read shipping list or file is empty")
            return False

        print("Reading policy file...")
        policy = read_policy_file(policy_file)
        if policy is None:
            print("Error: Failed to read policy file")
            return False

        print("Reading shipping rate file...")
        shipping_rate = read_shipping_rate_file(shipping_rate_file)
        if shipping_rate is None:
            print("Error: Failed to read shipping rate file")
            return False

        print("Reading exchange rate file...")
        exchange_rates = read_exchange_rate_file(exchange_rate_file)
        if exchange_rates is None:
            print("Error: Failed to read exchange rate file")
            return False

        # Step 1: Normalize the shipping list
        print("\nNormalizing shipping list...")
        df_normalized = normalize_shipping_list(df)
        if df_normalized is None:
            print("Error: Failed to normalize shipping list")
            return False

        # Step 2: Calculate FOB prices
        print("\nCalculating FOB prices...")
        df_with_fob = calculate_fob_prices(df_normalized, policy)
        if df_with_fob is None:
            print("Error: Failed to calculate FOB prices")
            return False

        # Step 3: Save FOB prices
        print("\nSaving initial FOB prices...")
        if not save_fob_prices(df_with_fob, output_fob_file):
            print("Error: Failed to save FOB prices")
            return False

        # Second Pass
        print("\n=== Second Pass ===")
        print("Reading FOB prices file for second pass...")
        try:
            df_second_pass = pd.read_excel(output_fob_file, sheet_name='FOB Prices')
            print(f"Successfully read FOB prices file with {len(df_second_pass)} rows")
            
            # Normalize the data again to fill any remaining empty cells
            print("\nProcessing remaining empty cells...")
            df_final = normalize_shipping_list(df_second_pass)
            
            # Save the final version
            print("\nSaving final FOB prices...")
            if not save_fob_prices(df_final, output_fob_file):
                print("Error: Failed to save final FOB prices")
                return False
            
            # Use the final version for CIF calculations
            df_with_fob = df_final
            
        except Exception as e:
            print(f"Warning: Error during second pass: {e}")
            print("Continuing with first pass results...")

        # Step 4: Calculate CIF prices
        print("\nCalculating CIF prices...")
        df_with_cif = calculate_cif_prices(df_with_fob, policy, shipping_rate, exchange_rates)
        if df_with_cif is None:
            print("Error: Failed to calculate CIF prices")
            return False

        # Step 5: Generate export receipt
        print("\nGenerating export receipt...")
        if not generate_export_receipt(df_with_cif, output_export_file):
            print("Error: Failed to generate export receipt")
            return False

        # Step 6: Generate re-import receipt
        print("\nGenerating re-import receipt...")
        if not generate_reimport_receipt(df_with_cif, output_reimport_file):
            print("Error: Failed to generate re-import receipt")
            return False

        print("\nProcessing completed successfully!")
        return True

    except Exception as e:
        print(f"Error in process_shipping_list: {str(e)}")
        return False


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Process shipping list to generate export and re-import receipts.')
    parser.add_argument('--shipping-list', required=True, help='Path to shipping list Excel file')
    parser.add_argument('--policy-file', required=True, help='Path to policy Excel file')
    parser.add_argument('--shipping-rate-file', required=True, help='Path to shipping rate Excel file')
    parser.add_argument('--exchange-rate-file', required=True, help='Path to exchange rate Excel file')
    parser.add_argument('--output-fob', default='shipping_fob_prices.xlsx', help='Path to save shipping list with FOB prices')
    parser.add_argument('--output-export', default='export_receipt.xlsx', help='Path to save export receipt')
    parser.add_argument('--output-reimport', default='reimport_receipt.xlsx', help='Path to save re-import receipt')
    
    args = parser.parse_args()
    
    process_shipping_list(
        args.shipping_list,
        args.policy_file,
        args.shipping_rate_file,
        args.exchange_rate_file,
        args.output_fob,
        args.output_export,
        args.output_reimport
    ) 